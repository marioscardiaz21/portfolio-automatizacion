REPORTE DE FALLA TECNICA
"""
✓ Fechas: evita ambigüedad DD/MM/YYYY vs YYYY-MM-DD y selecciona la más reciente real.
✓ DERIVADO = SI/NO según cruce con TRANSFERENCIAS (indicator).
✓ Conserva DERIVADO (ORIG) si existe.
✓ Homologación de canal por MOTIVO usando “HOMOLOGACION TRANSFERENCIAS”.
✓ Renombrado y orden EXACTO de columnas.
✓ Rellena “NO APLICA” en CANAL_DERIVADO y MOTIVO_DE_LA_DERIVACION tras el renombrado.
✓ Ignora archivos temporales “~$”.
"""

import os
import re
import unicodedata
import sys
import pandas as pd
from datetime import date, timedelta

# =========================
# ========= CONFIG ========
# =========================

BASE_DIR = r"\\SERVIDOR_RED\RUTA\PROYECTO\KPIS\FALLA TECNICA"

# Patrones: se toma el archivo MÁS RECIENTE que contenga el patrón en el nombre.
PATRONES_IBR = [
    ("IBR_KPI_BASE_CERO_PAGOS_",    "IBR"),
    ("IBR_KPI_BASE_FALLA_TECNICA_", "IBR"),
    ("IBR_KPI_BASE_UN_PAGO_",       "IBR"),
]
PATRON_TRANSF = ("TRANSFERENCIAS_", "TRANSF")

# Archivo de homologación
PATRON_HOMOLOG = "HOMOLOGACION TRANSFERENCIAS"
CAND_HOJA_HOMOLOG = ["Hoja1", "Sheet1", "HOMOLOGACION", "MAPA", "LISTA"]

# Hojas candidatas
CAND_HOJA_IBR = [
    "MOTIVO DE NO PAGO", "MOTIVOS DE NO PAGO", "MOTIVO NO PAGO",
    "MOTIVO_DE_NO_PAGO", "MOTIVO_NO_PAGO"
]
CAND_HOJA_TRANSF = ["Sheet1", "Hoja1", "TRANSFERENCIAS", "BASE", "DATA"]

# Columnas esperadas en el IBR unificado (etiquetas lógicas)
COLS_IBR_ESPERADAS = [
    "CUSTOMER_ID", "RUC_DNI", "GESTOR", "ETAPA", "PLAN", "GESTIONADO",
    "TIPO DE CONTACTO", "ESCENARIO DE TIPIFICACIÓN", "MOTIVO DE NO PAGO",
    "FECHA ULT GESTION", "PDP", "CLIENTES RECUPERADOS", "INTENSIDAD", "ESTADO_SERVICIO"
]

# Columnas candidatas en TRANSFERENCIAS
CAND_CUENTA_B = ["CUENTA", "CUSTOMER_ID", "CUST ACCOUNT", "CUST-ACCOUNT", "CUST_ACCOUNT"]
CAND_FECHA_B  = ["FECHA", "FECHA TRANSFERENCIA", "FECHA_GESTION", "FECHA ULT GESTION",
                 "FECHA_HORA", "HORA INICIO", "START_DATE"]
CAND_TIPO_TRANSF = [
    "TIPO_TRANSFERENCIA", "TIPO TRANSFERENCIA", "TIPO DE TRANSFERENCIA", "TIPO_DE_TRANSFERENCIA",
    "TIPO TRANSFERENCIAS", "TIPO DE TRANSFERENCIAS"
]
CAND_MOTIVO = ["MOTIVO", "MOTIVO DERIVACION", "MOTIVO DE LA DERIVACION", "MOTIVO DE LA DERIVACIÓN"]
CAND_DERIVADO = ["DERIVADO"]

JOIN_TYPE = "left"

# Salida
OUT_BASE_FOLDER_NAME = "FALLA DE SERVICIO"
SHEET_NAME = "FALLA DE SERVICIO"

# Formato Excel
HEADER_YELLOW = "FFFF00"   # primeras 6 columnas
HEADER_GREEN  = "2E7D32"   # resto
HEADER_FONT   = "FFFFFF"
DATE_FMT      = "DD/MM/YYYY"


# ======================================
# =========== UTILIDADES  ==============
# ======================================

def quitar_acentos(s):
    if s is None:
        return ""
    return ''.join(c for c in unicodedata.normalize('NFKD', str(s))
                   if not unicodedata.combining(c))

def normalizar_texto(s: str) -> str:
    s2 = quitar_acentos(s or "").strip().upper()
    s2 = re.sub(r"\s+", " ", s2)
    return s2

def resolver_hoja_por_candidatos(ruta, candidatos):
    xls = pd.ExcelFile(ruta, engine="openpyxl")
    hojas = xls.sheet_names
    norm = {normalizar_texto(h): h for h in hojas}

    # match exacto
    for cand in candidatos:
        c = normalizar_texto(cand)
        if c in norm:
            return norm[c]
    # match contiene
    for cand in candidatos:
        c = normalizar_texto(cand)
        for h_norm, real in norm.items():
            if c in h_norm:
                return real
    # fallback primera hoja
    return hojas[0]

def encontrar_archivo_mas_reciente_ci(base_dir, patron):
    patron = patron.lower()
    exts = (".xlsx", ".xlsm", ".xls", ".csv")
    archivos = [
        os.path.join(base_dir, f)
        for f in os.listdir(base_dir)
        if os.path.splitext(f)[1].lower() in exts
        and patron in f.lower()
        and not f.startswith("~$")  # ignorar temporales de Excel
    ]
    if not archivos:
        raise FileNotFoundError(f"No hay archivos para patrón '{patron}' en {base_dir}")
    archivos.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return archivos[0]

def leer_archivo(ruta, hoja=None):
    ext = os.path.splitext(ruta)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(ruta, dtype=str, sheet_name=hoja, engine="openpyxl")
    if ext == ".xls":
        return pd.read_excel(ruta, dtype=str, sheet_name=hoja, engine="xlrd")
    if ext == ".csv":
        return pd.read_csv(ruta, dtype=str)
    raise ValueError("Extensión no soportada")

def encontrar_col(df, candidatos):
    df_map = {normalizar_texto(c): c for c in df.columns}
    for cand in candidatos:
        key = normalizar_texto(cand)
        if key in df_map:
            return df_map[key]
    return None

def norm_key(x):
    if pd.isna(x): return ""
    s = str(x).strip().upper()
    return re.sub(r"[^0-9A-Z]", "", s)


# ===============================================================
# ===============  PARSEO DE FECHA CORREGIDO  ===================
# ===============================================================

def parse_fecha_ibr(serie):
    # Uniformizar separadores para evitar ambigüedad
    s = (
        serie.astype(str)
        .str.strip()
        .str.replace("-", "/", regex=False)     # 2026-02-12 -> 2026/02/12
    )
    # Intento fijo DD/MM/YYYY
    dt = pd.to_datetime(s, format="%d/%m/%Y", errors="coerce")
    # Intento YYYY/MM/DD para los que fallaron
    mask = dt.isna()
    if mask.any():
        dt2 = pd.to_datetime(s[mask], format="%Y/%m/%d", errors="coerce")
        dt.loc[mask] = dt2
    return dt.dt.date

def parse_fecha_b(serie):
    s = serie.astype(str).str.strip().str.replace("-", "/", regex=False)
    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    return dt.dt.date


# ===============================================================
# =================== HOMOLOGACIÓN CANAL  =======================
# ===============================================================

def leer_mapa_homologacion(base_dir):
    ruta = encontrar_archivo_mas_reciente_ci(base_dir, PATRON_HOMOLOG)
    hoja = resolver_hoja_por_candidatos(ruta, CAND_HOJA_HOMOLOG)
    df = pd.read_excel(ruta, dtype=str, sheet_name=hoja, header=None, engine="openpyxl")
    df = df.iloc[:, :2]
    df.columns = ["MOTIVO_SRC", "CANAL_DST"]
    df = df.dropna()
    df["KEY"] = df["MOTIVO_SRC"].map(normalizar_texto)
    return dict(zip(df["KEY"], df["CANAL_DST"])), ruta, hoja


# ===============================================================
# =========================== MAIN ==============================
# ===============================================================

def main():

    # -----------------------------------------
    # 1) Cargar y unificar IBR
    # -----------------------------------------
    partes = []
    for patron, _ in PATRONES_IBR:
        ruta = encontrar_archivo_mas_reciente_ci(BASE_DIR, patron)
        hoja = resolver_hoja_por_candidatos(ruta, CAND_HOJA_IBR)
        print(f"📂 Cargando IBR: {os.path.basename(ruta)} | Hoja='{hoja}'")
        df = leer_archivo(ruta, hoja)
        partes.append(df)

    df_ibr = pd.concat(partes, ignore_index=True)

    # Normalizar columnas a etiquetas esperadas
    mapa_norm = {normalizar_texto(c): c for c in df_ibr.columns}
    ren = {}
    for target in COLS_IBR_ESPERADAS:
        nt = normalizar_texto(target)
        if nt in mapa_norm:
            ren[mapa_norm[nt]] = target
    df_ibr = df_ibr.rename(columns=ren)

    # Asegurar columnas
    for col in COLS_IBR_ESPERADAS:
        if col not in df_ibr.columns:
            df_ibr[col] = pd.NA

    # FECHA: más reciente válida
    df_ibr["FECHA_DIA"] = parse_fecha_ibr(df_ibr["FECHA ULT GESTION"])
    hoy = date.today()
    fecha_min = date(2015, 1, 1)
    fecha_max = hoy + timedelta(days=1)

    df_ibr["FECHA_DIA"] = df_ibr["FECHA_DIA"].apply(
        lambda d: d if (pd.notna(d) and fecha_min <= d <= fecha_max) else pd.NA
    )

    fechas = [d for d in df_ibr["FECHA_DIA"] if pd.notna(d)]
    if not fechas:
        raise ValueError("No hay fechas válidas en IBR")
    fecha_dia = max(fechas)
    print(f"🗓️ Fecha final seleccionada: {fecha_dia}")

    antes = len(df_ibr)
    df_ibr = df_ibr[df_ibr["FECHA_DIA"] == fecha_dia].copy()
    print(f"✅ IBR filtrado: {antes} → {len(df_ibr)} filas")

    # -----------------------------------------
    # 2) Transferencias
    # -----------------------------------------
    ruta_b = encontrar_archivo_mas_reciente_ci(BASE_DIR, PATRON_TRANSF[0])
    hoja_b = resolver_hoja_por_candidatos(ruta_b, CAND_HOJA_TRANSF)
    print(f"📂 Cargando TRANSFERENCIAS: {os.path.basename(ruta_b)} | Hoja='{hoja_b}'")

    df_b = leer_archivo(ruta_b, hoja_b)

    col_cuenta  = encontrar_col(df_b, CAND_CUENTA_B)
    if not col_cuenta:
        raise KeyError("No se encontró columna de CUENTA/CUSTOMER_ID en TRANSFERENCIAS")

    col_fecha_b = encontrar_col(df_b, CAND_FECHA_B)
    col_tipo    = encontrar_col(df_b, CAND_TIPO_TRANSF)
    col_motivo  = encontrar_col(df_b, CAND_MOTIVO)
    col_deriv   = encontrar_col(df_b, CAND_DERIVADO)

    if col_fecha_b:
        df_b["FECHA_DIA"] = parse_fecha_b(df_b[col_fecha_b])
        antes_b = len(df_b)
        df_b = df_b[df_b["FECHA_DIA"] == fecha_dia].copy()
        print(f"✓ TRANSFERENCIAS filtrado: {antes_b} → {len(df_b)}")
    else:
        print("⚠️ TRANSFERENCIAS sin columna de fecha; se cruza sin filtrar por fecha.")

    # -----------------------------------------
    # 3) Merge con indicador
    # -----------------------------------------
    df_ibr["_KEY"] = df_ibr["CUSTOMER_ID"].astype(str).apply(norm_key)
    df_b["_KEY"]   = df_b[col_cuenta].astype(str).apply(norm_key)

    cols_keep_b = [c for c in [col_tipo, col_motivo, col_deriv] if c]
    merged = df_ibr.merge(
        df_b[["_KEY"] + cols_keep_b],
        how=JOIN_TYPE,
        on="_KEY",
        indicator=True
    )

    # -----------------------------------------
    # 4) Construcción salida (nombres originales)
    # -----------------------------------------
    derivado_flag = merged["_merge"].eq("both").map({True: "SI", False: "NO"})
    derivado_orig = merged[col_deriv] if col_deriv else pd.Series([pd.NA]*len(merged))

    out = pd.DataFrame({
        "TIPO DE FOCALIZADO": ["FALLA TECNICA"]*len(merged),
        "RUC_DNI": merged["RUC_DNI"],
        "CUSTOMER_ID": merged["CUSTOMER_ID"],
        "GESTOR": merged["GESTOR"],
        "ETAPA": merged["ETAPA"],
        "PLAN": merged["PLAN"],
        "GESTIONADO": merged["GESTIONADO"],
        "TIPO DE CONTACTO": merged["TIPO DE CONTACTO"],
        "ESCENARIO DE TIPIFICACIÓN": merged["ESCENARIO DE TIPIFICACIÓN"],
        "MOTIVOS DE NO PAGO": merged["MOTIVO DE NO PAGO"],
        "FECHA ULT GESTION": merged["FECHA_DIA"],
        "DERIVADO": derivado_flag,
        "CANAL DERIVADO": merged[col_tipo] if col_tipo else pd.NA,
        "MOTIVO DE LA DERIVACIÓN": merged[col_motivo] if col_motivo else pd.NA,
        "PDP": merged["PDP"],
        "CLIENTES RECUPERADOS": merged["CLIENTES RECUPERADOS"],
        "INTENSIDAD": merged["INTENSIDAD"],
        "ESTADO DEL SERVICIO": merged["ESTADO_SERVICIO"]
    })

    if col_deriv:
        out.insert(out.columns.get_loc("DERIVADO")+1, "DERIVADO (ORIG)", derivado_orig)

    # -----------------------------------------
    # 5) Homologación (se impone canal por motivo si hay mapeo)
    # -----------------------------------------
    try:
        mapa, ruta_h, hoja_h = leer_mapa_homologacion(BASE_DIR)
        keys = out["MOTIVO DE LA DERIVACIÓN"].map(normalizar_texto)
        mapped = keys.map(mapa)
        out["CANAL DERIVADO"] = mapped.fillna(out["CANAL DERIVADO"])
        # Nota: el relleno de NO APLICA lo haremos luego del renombrado
        print("Homologación aplicada.")
    except Exception as e:
        print("⚠️ No se pudo aplicar homologación:", e)

    # -----------------------------------------
    # 6) Renombrado final + ORDEN EXACTO + Relleno “NO APLICA”
    # -----------------------------------------
    rename_map = {
        "TIPO DE FOCALIZADO": "TIPO_FOCALIZADO",
        "RUC_DNI": "DOCUMENTO",
        "ESCENARIO DE TIPIFICACIÓN": "ESCENARIOS_TIPIFICACION",
        "MOTIVOS DE NO PAGO": "MOTIVO_NO_PAGO",
        "FECHA ULT GESTION": "FECHA_ULT_GESTION",
        "CANAL DERIVADO": "CANAL_DERIVADO",
        "MOTIVO DE LA DERIVACIÓN": "MOTIVO_DE_LA_DERIVACION",
        "CLIENTES RECUPERADOS": "CLIENTE_RECUPERADO",
        "ESTADO DEL SERVICIO": "ESTADO_SERVICIO",
    }
    out = out.rename(columns=rename_map)

    # === Rellenar con "NO APLICA" tras el renombrado ===
    for col in ["CANAL_DERIVADO", "MOTIVO_DE_LA_DERIVACION"]:
        if col in out.columns:
            out[col] = (
                out[col]
                .astype(str)
                .str.strip()
                .replace({"": "NO APLICA", "nan": "NO APLICA", "None": "NO APLICA"})
                .fillna("NO APLICA")
            )

    # ORDEN EXACTO solicitado
    orden_final = [
        "TIPO_FOCALIZADO",
        "DOCUMENTO",
        "CUSTOMER_ID",
        "GESTOR",
        "ETAPA",
        "PLAN",
        "GESTIONADO",
        "TIPO DE CONTACTO",
        "ESCENARIOS_TIPIFICACION",
        "MOTIVO_NO_PAGO",
        "FECHA_ULT_GESTION",
        "DERIVADO",
        "CANAL_DERIVADO",
        "MOTIVO_DE_LA_DERIVACION",
        "PDP",
        "CLIENTE_RECUPERADO",
        "INTENSIDAD",
        "ESTADO_SERVICIO",
    ]
    if "DERIVADO (ORIG)" in out.columns:
        idx = orden_final.index("DERIVADO") + 1
        orden_final.insert(idx, "DERIVADO (ORIG)")

    out = out[[c for c in orden_final if c in out.columns]]

    # -----------------------------------------
    # 7) Exportar Excel con formato
    # -----------------------------------------
    fecha_folder = fecha_dia.strftime("%Y%m%d")
    out_dir = os.path.join(BASE_DIR, OUT_BASE_FOLDER_NAME, fecha_folder)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"FALLA DE SERVICIO_{fecha_folder}.xlsx")

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        ws = writer.sheets[SHEET_NAME]

        ws.freeze_panes = "A2"

        # Encabezado con dos colores
        for j, cell in enumerate(ws[1], start=1):
            cell.font = Font(color=HEADER_FONT, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if j <= 6:
                cell.fill = PatternFill(start_color=HEADER_YELLOW, end_color=HEADER_YELLOW, fill_type="solid")
            else:
                cell.fill = PatternFill(start_color=HEADER_GREEN, end_color=HEADER_GREEN, fill_type="solid")

        # Auto-anchos aproximados
        for col_idx, col in enumerate(out.columns, start=1):
            max_len = max([len(str(col))] + [len(str(x)) for x in out[col].astype(str).head(800)])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        # Formato fecha robusto (tolera ambos nombres si cambian)
        fecha_col_name = None
        for cand in ["FECHA_ULT_GESTION", "FECHA ULT GESTION"]:
            if cand in out.columns:
                fecha_col_name = cand
                break
        if fecha_col_name:
            idx_fecha = list(out.columns).index(fecha_col_name) + 1
            for row in ws.iter_rows(min_row=2, min_col=idx_fecha, max_col=idx_fecha, max_row=ws.max_row):
                for cell in row:
                    cell.number_format = DATE_FMT

    print("✅ Archivo generado:", out_path)


# ===============================================================
# ========================= EJECUCIÓN ===========================
# ===============================================================

if __name__ == "__main__":
    main()
    if not os.environ.get("PYCHARM_HOSTED") and sys.stdin.isatty():
        input("\nPresiona Enter para cerrar...")

